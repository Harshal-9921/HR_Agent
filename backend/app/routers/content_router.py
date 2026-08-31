from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from .logs_router import log_action
from typing import List
import os
import shutil
from datetime import datetime
from .. import models, schemas, auth
from ..database import get_db
from fastapi.responses import StreamingResponse
import csv
import io
try:
    import openpyxl
except ImportError:
    openpyxl = None

router = APIRouter(prefix="/api/content", tags=["Content Management"])


UPLOAD_DIR = "static/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# --- Content Endpoints ---

@router.get("/", response_model=List[schemas.ContentResponse])
def get_all_content(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    all_content = db.query(models.Content).order_by(models.Content.order).all()
    if current_user.role in [models.RoleEnum.hr, models.RoleEnum.admin]:
        return all_content
    visible = []
    for c in all_content:
        dept_ok = True
        if c.visible_departments:
            allowed_depts = [d.strip() for d in c.visible_departments.split(",")]
            dept_ok = current_user.department in allowed_depts

        role_ok = True
        if c.role_visibility:
            allowed_roles = [r.strip() for r in c.role_visibility.split(",")]
            role_ok = current_user.role.value in allowed_roles

        if dept_ok and role_ok:
            visible.append(c)
    return visible

@router.post("/", response_model=schemas.ContentResponse)
def create_content(
    data: schemas.ContentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    new_content = models.Content(**data.dict())
    db.add(new_content)
    db.commit()
    db.refresh(new_content)
    return new_content

@router.put("/reorder")
def reorder_modules(
    order_data: list[dict],
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    """HR reorders non-intro modules. Expects list of {id, order}"""
    for item in order_data:
        content = db.query(models.Content).filter(
            models.Content.id == item["id"],
            models.Content.is_intro == False
        ).first()
        if content:
            content.order = item["order"]
    db.commit()
    return {"message": "Order updated successfully."}


@router.post("/complete-module", response_model=schemas.ModuleProgressResponse)
def complete_module(
    data: schemas.ModuleProgressCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    if current_user.role in [models.RoleEnum.hr, models.RoleEnum.admin]:
        raise HTTPException(status_code=403, detail="HR/Admin accounts cannot complete onboarding modules directly.")

    passed = data.total_questions == 0 or (data.score / data.total_questions) >= 0.5

    existing = db.query(models.ModuleProgress).filter(
        models.ModuleProgress.user_id == current_user.id,
        models.ModuleProgress.content_id == data.content_id
    ).first()

    if existing:
        if existing.completed:
            return existing

        if existing.attempt_count >= 2:
            raise HTTPException(
                status_code=403,
                detail="Maximum attempts reached. You have been reassigned to the module."
            )

        existing.attempt_count += 1
        existing.score = data.score
        existing.total_questions = data.total_questions

        if passed:
            existing.completed = True
            existing.completed_at = datetime.now().isoformat()

        db.commit()
        if passed:
            log_action(db, current_user.id, "MODULE_COMPLETED", f"Completed module ID: {data.content_id} with score {data.score}/{data.total_questions}")
        db.refresh(existing)

        if passed:
            all_content_count = db.query(models.Content).filter(models.Content.is_keka == False).count()
            completed_count = db.query(models.ModuleProgress).filter(
                models.ModuleProgress.user_id == current_user.id,
                models.ModuleProgress.completed == True
            ).count()
            if all_content_count > 0 and completed_count >= all_content_count:
                from ..worker import send_completion_email_to_hr
                send_completion_email_to_hr.delay(current_user.id)

        if not passed and existing.attempt_count >= 2:
            raise HTTPException(
                status_code=403,
                detail="Maximum attempts reached. You have been reassigned to the module."
            )

        return existing

    else:
        new_progress = models.ModuleProgress(
            user_id=current_user.id,
            content_id=data.content_id,
            completed=passed,
            score=data.score,
            total_questions=data.total_questions,
            attempt_count=1,
            completed_at=datetime.now().isoformat() if passed else None
        )
        db.add(new_progress)

        if passed:
            onboarding = db.query(models.OnboardingProgress).filter(
                models.OnboardingProgress.user_id == current_user.id
            ).first()
            if onboarding:
                all_content_count = db.query(models.Content).filter(models.Content.is_keka == False).count()
                completed_count = db.query(models.ModuleProgress).filter(
                    models.ModuleProgress.user_id == current_user.id,
                    models.ModuleProgress.completed == True
                ).count() + 1
                if all_content_count > 0:
                    onboarding.completion_percentage = int((completed_count / all_content_count) * 100)
                onboarding.last_activity_at = datetime.now().isoformat()

        db.commit()
        db.refresh(new_progress)

        if passed:
            all_content_count = db.query(models.Content).count()
            completed_count = db.query(models.ModuleProgress).filter(
                models.ModuleProgress.user_id == current_user.id,
                models.ModuleProgress.completed == True
            ).count()
            if all_content_count > 0 and completed_count >= all_content_count:
                from ..worker import send_completion_email_to_hr
                send_completion_email_to_hr.delay(current_user.id)

        if not passed and new_progress.attempt_count >= 2:
            raise HTTPException(
                status_code=403,
                detail="Maximum attempts reached. You have been reassigned to the module."
            )

        return new_progress

# --- MCQ Endpoints ---

@router.get("/{content_id}/mcqs", response_model=List[schemas.MCQResponse])
def get_mcqs_for_content(content_id: int, db: Session = Depends(get_db)):
    return db.query(models.MCQ).filter(models.MCQ.content_id == content_id).all()

@router.post("/mcqs", response_model=schemas.MCQResponse)
def create_mcq(
    data: schemas.MCQCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    parent = db.query(models.Content).filter(models.Content.id == data.content_id).first()
    if parent and parent.is_keka:
        raise HTTPException(status_code=400, detail="MCQs cannot be added to the Keka Acknowledgement step")
    new_mcq = models.MCQ(**data.dict())
    db.add(new_mcq)
    db.commit()
    db.refresh(new_mcq)
    return new_mcq
@router.put("/mcqs/{mcq_id}", response_model=schemas.MCQResponse)
def update_mcq(
    mcq_id: int,
    data: schemas.MCQCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    mcq = db.query(models.MCQ).filter(models.MCQ.id == mcq_id).first()
    if not mcq:
        raise HTTPException(status_code=404, detail="MCQ not found")
    for key, value in data.dict().items():
        setattr(mcq, key, value)
    db.commit()
    db.refresh(mcq)
    return mcq

@router.delete("/mcqs/{mcq_id}")
def delete_mcq(
    mcq_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    mcq = db.query(models.MCQ).filter(models.MCQ.id == mcq_id).first()
    if not mcq:
        raise HTTPException(status_code=404, detail="MCQ not found")
    
    db.delete(mcq)
    db.commit()
    return {"message": "MCQ deleted"}

@router.post("/{content_id}/mcqs/bulk")
async def bulk_upload_mcqs(
    content_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    content = db.query(models.Content).filter(models.Content.id == content_id).first()
    if not content:
        raise HTTPException(status_code=404, detail="Content module not found")
    if content.is_keka:
        raise HTTPException(status_code=400, detail="MCQs cannot be added to the Keka Acknowledgement step")

    filename = file.filename or ""
    file_bytes = await file.read()

    rows = []
    if filename.lower().endswith(".csv"):
        text = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append(r)
    elif filename.lower().endswith((".xlsx", ".xls")):
        if not openpyxl:
            raise HTTPException(status_code=500, detail="Excel support not installed on server")
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append(dict(zip(headers, row)))
    else:
        raise HTTPException(status_code=400, detail="Unsupported file type. Please upload .csv or .xlsx")

    created = 0
    errors = []
    for idx, row in enumerate(rows, start=2):
        try:
            question = str(row.get("Question") or row.get("question") or "").strip()
            option_a = str(row.get("Option A") or row.get("option_a") or "").strip()
            option_b = str(row.get("Option B") or row.get("option_b") or "").strip()
            option_c = str(row.get("Option C") or row.get("option_c") or "").strip()
            option_d = str(row.get("Option D") or row.get("option_d") or "").strip()
            correct = str(row.get("Correct Answer") or row.get("correct_answer") or "").strip().upper()

            if not question or not option_a or not option_b or not option_c or not option_d:
                errors.append(f"Row {idx}: missing required field(s), skipped")
                continue
            if correct not in ("A", "B", "C", "D"):
                errors.append(f"Row {idx}: invalid Correct Answer '{correct}' (must be A/B/C/D), skipped")
                continue

            mcq = models.MCQ(
                content_id=content_id,
                question=question,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                correct_answer=correct
            )
            db.add(mcq)
            created += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    db.commit()
    return {"created": created, "errors": errors}


@router.get("/mcqs/template/csv")
def download_mcq_template_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Question", "Option A", "Option B", "Option C", "Option D", "Correct Answer"])
    writer.writerow(["What is the capital of France?", "Paris", "London", "Berlin", "Madrid", "A"])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=mcq_upload_template.csv"}
    )


@router.get("/mcqs/template/xlsx")
def download_mcq_template_xlsx():
    if not openpyxl:
        raise HTTPException(status_code=500, detail="Excel support not installed on server")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Question", "Option A", "Option B", "Option C", "Option D", "Correct Answer"])
    ws.append(["What is the capital of France?", "Paris", "London", "Berlin", "Madrid", "A"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mcq_upload_template.xlsx"}
    )

# --- File Upload ---

@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    file_path = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Return the URL path
    return {"url": f"/static/uploads/{file.filename}"}

# --- Module Progress ---

@router.get("/my-progress", response_model=List[schemas.ModuleProgressResponse])
def get_my_progress(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """Get all completed modules for the current user."""
    return db.query(models.ModuleProgress).filter(
        models.ModuleProgress.user_id == current_user.id
    ).all()

@router.get("/employee-progress/{user_id}", response_model=List[schemas.ModuleProgressResponse])
def get_employee_progress(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    """HR: Get module progress for a specific employee."""
    return db.query(models.ModuleProgress).filter(
        models.ModuleProgress.user_id == user_id
    ).all()

@router.post("/reset-attempts/{content_id}")
def reset_attempts(
    content_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    """Reset attempt count when employee restarts the module."""
    existing = db.query(models.ModuleProgress).filter(
        models.ModuleProgress.user_id == current_user.id,
        models.ModuleProgress.content_id == content_id
    ).first()
    if existing:
        existing.attempt_count = 0
        existing.completed = False
        existing.score = 0
        db.commit()
    return {"message": "Attempts reset successfully."}

# --- Content Update ---

@router.put("/{content_id}", response_model=schemas.ContentResponse)
def update_content(
    content_id: int,
    data: schemas.ContentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    content = db.query(models.Content).filter(models.Content.id == content_id).first()
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    for key, value in data.dict().items():
        if key == 'is_intro':
            continue  
        setattr(content, key, value)
    if content.is_intro:
        content.order = 0
    db.commit()
    db.refresh(content)
    return content
    content = db.query(models.Content).filter(
        models.Content.id == content_id
    ).first()

    if not content:
        raise HTTPException(status_code=404, detail="Content not found")

    content.title = data.title
    content.description = data.description
    content.content_type = data.content_type
    content.file_url = data.file_url

    db.commit()
    db.refresh(content)

    return content


# --- Content Delete ---

@router.delete("/{content_id}")
def delete_content(
    content_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(
        auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin])
    )
):
    content = db.query(models.Content).filter(
        models.Content.id == content_id
    ).first()

    if not content:
        raise HTTPException(status_code=404, detail="Content not found")

    if content.is_intro:
        raise HTTPException(
            status_code=400,
            detail="Introduction module cannot be deleted"
        )
    if content.is_keka:
        raise HTTPException(
            status_code=400,
            detail="Keka Acknowledgement cannot be deleted"
        )

    db.delete(content)
    db.commit()

    return {"message": "Content deleted successfully"}

@router.post("/reorder")
def reorder_content(
    data: schemas.ReorderRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(
        auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin])
    )
):
    intro = db.query(models.Content).filter(
        models.Content.is_intro == True
    ).first()

    if intro:
        intro.order = 0

    for index, content_id in enumerate(data.content_ids, start=1):
        content = db.query(models.Content).filter(
            models.Content.id == content_id
        ).first()

        if content and not content.is_intro:
            content.order = index

    db.commit()

    return {"message": "Order updated successfully"}

@router.put("/{content_id}/toggle")
def toggle_module(
    content_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.require_role([models.RoleEnum.hr, models.RoleEnum.admin]))
):
    """HR enables or disables a module."""
    content = db.query(models.Content).filter(models.Content.id == content_id).first()
    if not content:
        raise HTTPException(status_code=404, detail="Content not found")
    if content.is_intro:
        raise HTTPException(status_code=400, detail="Introduction module cannot be disabled.")
    content.is_enabled = not content.is_enabled
    db.commit()
    db.refresh(content)
    return {"id": content.id, "is_enabled": content.is_enabled}